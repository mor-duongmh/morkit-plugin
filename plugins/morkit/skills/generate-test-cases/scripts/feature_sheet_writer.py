"""Select/clone a Feature sheet and write test-case rows into it.

Column map (per inspected template): A=ID(formula) B=Description C=Pre-Condition
D=Procedure E=Expected F=R0 G=R1 H=R2 I=R3 J=Final(formula) K=BugID L=Note.
Data region starts at row 9; row 8 is the first section header. We fill only
B/C/D/E + section (B-only) rows, and stamp A/J formulas. We never write F-I/K/L.
"""

from __future__ import annotations

import copy
import re

from openpyxl.worksheet.datavalidation import DataValidation

PLACEHOLDER_RE = re.compile(r"^Feature \d+$")   # template's placeholder tab/C1 name
COUNT_SIG = "=COUNT(A9"     # C2 signature that marks a feature sheet (name-independent)
INVALID_TITLE = re.compile(r"[:\\/?*\[\]]")     # chars Excel forbids in a sheet title
MAX_TITLE = 31              # Excel sheet-title length limit
DATA_START_ROW = 9          # first case row in the template
FIRST_SECTION_ROW = 8       # template's first section header row
PROTO_ROW = 9               # style prototype row
DV_LIST = '"Pass,Fail,Untested,N/A"'

A_FORMULA = '=IF($E{r}<>"",COUNTA($E$9:$E{r}),"")'
J_FORMULA = '=IF(I{r}<>"",I{r},IF(H{r}<>"",H{r},G{r}))'


def _style_of(ws, row, col):
    c = ws.cell(row, col)
    return {
        "font": copy.copy(c.font),
        "border": copy.copy(c.border),
        "fill": copy.copy(c.fill),
        "alignment": copy.copy(c.alignment),
        "number_format": c.number_format,
    }


def _apply(cell, proto, wrap=False):
    cell.font = copy.copy(proto["font"])
    cell.border = copy.copy(proto["border"])
    cell.fill = copy.copy(proto["fill"])
    al = copy.copy(proto["alignment"])
    if wrap:
        al.wrap_text = True
    cell.alignment = al
    cell.number_format = proto["number_format"]


def _has_real_cases(ws) -> bool:
    """True if the sheet already holds written test cases (non-placeholder)."""
    for r in range(DATA_START_ROW, ws.max_row + 1):
        e = ws.cell(r, 5).value
        if e and not str(e).startswith("Sample output"):
            return True
    return False


def _last_data_row(ws) -> int:
    """Highest row whose Expected (E) is filled; DATA_START_ROW-1 if none."""
    last = DATA_START_ROW - 1
    for r in range(DATA_START_ROW, ws.max_row + 1):
        if ws.cell(r, 5).value not in (None, ""):
            last = r
    return last


def is_feature_sheet(ws) -> bool:
    """A feature sheet carries the template's `=COUNT(A9...)` formula in C2.

    Name-independent — survives tab renames so Test Report can re-discover
    every feature sheet by structure, not by a 'Feature N' label.
    """
    return str(ws["C2"].value or "").startswith(COUNT_SIG)


def feature_sheets(wb):
    """Feature sheets in workbook order."""
    return [ws for ws in wb.worksheets if is_feature_sheet(ws)]


def find_or_make_sheet(wb, feature_name: str, conflict: str):
    """Return (ws, is_new_clone, existed_for_feature)."""
    feats = feature_sheets(wb)

    # 1) sheet already used by this feature (C1 == feature_name).
    #    Skipped when conflict == 'new' so the feature lands on a fresh sheet.
    if conflict != "new":
        for ws in feats:
            if str(ws["C1"].value or "").strip() == feature_name:
                return ws, False, True

    # 2) an unused placeholder sheet (C1 still 'Feature N', no real cases)
    for ws in feats:
        c1 = str(ws["C1"].value or "").strip()
        if PLACEHOLDER_RE.match(c1) and not _has_real_cases(ws):
            return ws, False, False

    # 3) clone the last feature sheet as a prototype (temp title; renamed in finalize)
    if not feats:
        raise ValueError("Template has no feature sheet to clone from.")
    new = wb.copy_worksheet(feats[-1])
    new.title = _unique_title(wb, "Feature", keep=new)
    _reset_dv(new)  # copy_worksheet drops data-validations; re-add below anyway
    return new, True, False


def _sanitize_title(name: str) -> str:
    """Excel-legal sheet title: strip forbidden chars, collapse spaces, clamp len."""
    t = INVALID_TITLE.sub(" ", str(name)).strip()
    t = re.sub(r"\s+", " ", t)
    return (t or "Feature")[:MAX_TITLE]


def _unique_title(wb, name: str, keep=None) -> str:
    """A sanitized title unique among sheets (ignoring `keep`'s own current title)."""
    base = _sanitize_title(name)
    taken = {s for s in wb.sheetnames if keep is None or wb[s] is not keep}
    if base not in taken:
        return base
    n = 2
    while True:
        suffix = f" ({n})"
        cand = base[:MAX_TITLE - len(suffix)] + suffix
        if cand not in taken:
            return cand
        n += 1


def finalize_sheets(wb, active_ws, feature_name: str):
    """Rename the active feature tab to its real name and drop unused placeholders.

    Placeholder feature sheets (C1 still 'Feature N', no real cases) other than
    the active one are deleted, so the workbook keeps only tabs in actual use.
    The active sheet always survives and stays clonable as a future prototype.
    """
    active_ws.title = _unique_title(wb, feature_name, keep=active_ws)
    for ws in feature_sheets(wb):
        if ws is active_ws:
            continue
        c1 = str(ws["C1"].value or "").strip()
        if PLACEHOLDER_RE.match(c1) and not _has_real_cases(ws):
            del wb[ws.title]
    return active_ws.title


def _reset_dv(ws):
    ws.data_validations.dataValidation = []


def set_dv(ws, last_row: int):
    """Replace list-validations with one dropdown over F9:I{last_row}."""
    keep = [dv for dv in ws.data_validations.dataValidation
            if not (dv.type == "list" and "Pass" in str(dv.formula1))]
    ws.data_validations.dataValidation = keep
    if last_row < DATA_START_ROW:
        return
    dv = DataValidation(type="list", formula1=DV_LIST, allow_blank=True)
    dv.add(f"F{DATA_START_ROW}:I{last_row}")
    ws.add_data_validation(dv)


def clear_region(ws, upto: int):
    """Blank B..L values for rows FIRST_SECTION_ROW..upto (keep A formulas)."""
    for r in range(FIRST_SECTION_ROW, upto + 1):
        for col in range(2, 13):  # B..L
            ws.cell(r, col).value = None


def write_feature(ws, data: dict, append: bool):
    """Write C1 + sections/cases. Returns (last_row, row_count)."""
    section_proto = _style_of(ws, FIRST_SECTION_ROW, 2)
    data_proto = {c: _style_of(ws, PROTO_ROW, c) for c in range(1, 13)}

    ws["C1"] = data["feature"]

    if append:
        r = _last_data_row(ws) + 1
    else:
        clear_region(ws, max(40, ws.max_row))
        r = FIRST_SECTION_ROW

    count = 0
    for section in data["sections"]:
        ws.cell(r, 2).value = section["name"]
        _apply(ws.cell(r, 2), section_proto)
        r += 1
        for case in section["cases"]:
            for i, row in enumerate(case["rows"]):
                ws.cell(r, 1).value = A_FORMULA.format(r=r)
                _apply(ws.cell(r, 1), data_proto[1])
                desc = case["description"] if i == 0 else None
                _set(ws, r, 2, desc, data_proto[2])
                _set(ws, r, 3, row.get("precondition"), data_proto[3], wrap=True)
                _set(ws, r, 4, row.get("procedure"), data_proto[4], wrap=True)
                _set(ws, r, 5, row.get("expected"), data_proto[5], wrap=True)
                for col in (6, 7, 8, 9):  # F-I result cells: style only, no value
                    _apply(ws.cell(r, col), data_proto[col])
                ws.cell(r, 10).value = J_FORMULA.format(r=r)
                _apply(ws.cell(r, 10), data_proto[10])
                r += 1
                count += 1

    last_row = r - 1
    set_dv(ws, last_row)
    return last_row, count


def _set(ws, r, col, value, proto, wrap=False):
    cell = ws.cell(r, col)
    cell.value = value
    _apply(cell, proto, wrap=wrap)
